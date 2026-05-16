"""Importer: parse Closer Charts .xlsx + fetch 40-man pitcher rosters,
fuzzy-match existing chart names to MLBAMIDs, write JSON for the webapp.

Outputs (in --out, default ./data):
  rosters.json     - {teamId: {team, abbr, league, pitchers:[{id, name, posAbbr, hand, status}]}}
  chart.json       - {teams:[{teamId, league, levcon, roles:{closer:[chip], ...}, notes}]}
                     chip = {name, mlbamid, statusTag, color, other}
                     (usageTags are derived live from stats.json, never stored here)
  quickhits.json   - [{date, entries:{teamId: text}}]
  name_map.json    - {originalName: mlbamid|null}  (persisted for re-runs)

Usage:
  python3 build_data.py                 # interactive resolve for unmatched
  python3 build_data.py --auto          # auto-accept best guess; unresolved -> null
  python3 build_data.py --xlsx PATH     # custom workbook path
  python3 build_data.py --out DIR       # custom output dir
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl
from rapidfuzz import fuzz, process

# Local subset of mlbStats_26 endpoints (re-implemented without google.colab deps;
# same statsapi.mlb.com / bdfed URLs).
sys.path.insert(0, str(Path(__file__).resolve().parent))
import _mlb_api as mlb  # noqa: E402

# -----------------------------------------------------------------------------

DEFAULT_XLSX = Path(__file__).resolve().parent / "Copy of Closer Charts.xlsx"
DEFAULT_OUT = Path(__file__).resolve().parent / "data"

ROLE_COLS = {
    4: "closer",            # Closer (highest leverage)
    5: "stopper",           # Stopper (unless fluid)
    6: "stealth",           # Stealth Option
    7: "upside",            # High Upside Future
    8: "hybrid",            # Hybrid (IP+) or RP Pivot
}
LEVCON_COL = 2
TEAM_COL = 3
COLOR_COL = 1
NOTES_COL = 10

KNOWN_COLORS = {"Orange", "Yellow", "Magenta", "Green"}
TEAM_ROW_RANGE = range(4, 19)            # rows 4-18 inclusive (15 teams per league)
QUICK_HITS_START_ROW = 20

# Annotation parsing for chip cells:
#   - parenthetical "(paternity)"   -> statusTag
#   - square-bracket "[Blue]"       -> color (must match known palette)
#   - curly "{HLR}"                  -> other (free-text role/note)
#   - "3-of-4" / "1-of-2" / "Xx" "Xy" etc. -> drop (was previously captured as
#     `share`; the user clarified these denote recent usage, which we recompute
#     live from gameLog and store on stats.json, not on the chip).
SHARE_RE = re.compile(r"(\d+\s*-\s*of\s*-\s*\d+)", re.IGNORECASE)
PAREN_RE = re.compile(r"\(([^)]+)\)")
COLOR_RE = re.compile(r"\[([^\]]+)\]")
OTHER_RE = re.compile(r"\{([^}]+)\}")
COLOR_PALETTE = {"Orange", "Yellow", "Magenta", "Green", "Blue"}

# -----------------------------------------------------------------------------
# Excel parsing


def split_chip_text(raw: str) -> list[dict]:
    """Split a role-cell string into chip dicts.

    Examples:
      'Yennier Cano/Keegan Akin'              -> 2 chips (no annotations)
      'David Bednar 3-of-4'                   -> 1 chip (3-of-4 is dropped; usage
                                                 is recomputed from gameLog)
      'Graham Ashcraft (HLR)'                 -> 1 chip, statusTag='HLR'
      'Jose A. Ferrer (paternity)'            -> 1 chip, statusTag='paternity'
      'Mason Miller [Blue]'                   -> 1 chip, color='Blue'
      'Cade Smith {HLR}'                      -> 1 chip, other='HLR'
    """
    if not raw:
        return []
    raw = raw.strip()
    chips = []
    for piece in re.split(r"\s*/\s*", raw):
        piece = piece.strip()
        if not piece:
            continue

        # Drop share-style "3-of-4" tokens — they're recomputed live.
        piece = SHARE_RE.sub("", piece)

        # Pull bracketed color(s). New shape supports a comma-separated list
        # inside one set of brackets, e.g. "[Blue, Magenta]". Single legacy
        # "[Blue]" still parses. Unknown palette names are silently dropped.
        colors: list[str] = []
        cm = COLOR_RE.search(piece)
        if cm:
            raw_colors = cm.group(1)
            for token in re.split(r"\s*,\s*", raw_colors):
                cand = token.strip().capitalize()
                if cand in COLOR_PALETTE:
                    colors.append(cand)
            piece = piece[: cm.start()] + piece[cm.end():]

        # Pull curly "other" notes.
        other_vals: list[str] = []
        for om in OTHER_RE.finditer(piece):
            other_vals.append(om.group(1).strip())
        piece = OTHER_RE.sub("", piece)

        # Pull parenthetical status tags (paternity/IL/etc.)
        status_vals: list[str] = []
        for pm in PAREN_RE.finditer(piece):
            status_vals.append(pm.group(1).strip())
        piece = PAREN_RE.sub("", piece)

        name = re.sub(r"\s+", " ", piece).strip()
        if not name:
            continue
        chips.append({
            "name": name,
            "statusTag": status_vals[0] if status_vals else None,
            "colors": colors,
            "other": other_vals[0] if other_vals else None,
        })
    return chips


def parse_team_row(ws, row: int, league: str) -> dict | None:
    team = ws.cell(row=row, column=TEAM_COL).value
    if not team or not isinstance(team, str):
        return None
    team = team.strip()
    if not team:
        return None

    # Team-level color flag was removed from the UI model (colors live on
    # individual player chips now). We still ignore the legacy team color
    # column on re-import to keep the data clean.

    levcon_raw = ws.cell(row=row, column=LEVCON_COL).value
    try:
        levcon = int(levcon_raw) if levcon_raw is not None else None
    except (TypeError, ValueError):
        levcon = None

    roles: dict[str, list[dict]] = {}
    for col, role_key in ROLE_COLS.items():
        cell_val = ws.cell(row=row, column=col).value
        roles[role_key] = split_chip_text(cell_val if isinstance(cell_val, str) else "")

    notes_raw = ws.cell(row=row, column=NOTES_COL).value
    notes = notes_raw.strip() if isinstance(notes_raw, str) else ""

    return {
        "teamName": team,
        "league": league,
        "levcon": levcon,
        "roles": roles,
        "notes": notes,
    }


# Quick Hits per-team blurbs in the user's spreadsheet are written
# column-by-column, but the COLUMN position doesn't match the team header --
# the team is encoded as a "Cleveland: ..." / "Yankees: ..." prefix at the
# start of each blurb. At import time we (a) detect that prefix, (b) re-route
# the blurb under the matching team's entry, and (c) drop the redundant
# prefix from the display text.
QH_TEAM_PREFIX_RE = re.compile(r"^([A-Z][A-Za-z.\-' ]{0,30}?):\s+")

# Prefix aliases per league. Keys are lowercased.
_AL_ALIASES = {
    "baltimore": "Baltimore Orioles",  "orioles": "Baltimore Orioles",
    "boston": "Boston Red Sox",        "red sox": "Boston Red Sox",
    "chicago": "Chicago White Sox",    "white sox": "Chicago White Sox",  "cws": "Chicago White Sox",
    "cleveland": "Cleveland Guardians","guardians": "Cleveland Guardians",
    "detroit": "Detroit Tigers",       "tigers": "Detroit Tigers",
    "houston": "Houston Astros",       "astros": "Houston Astros",
    "kansas city": "Kansas City Royals","royals": "Kansas City Royals",  "kc": "Kansas City Royals",
    "angels": "Los Angeles Angels",    "laa": "Los Angeles Angels",       "halos": "Los Angeles Angels",
    "minnesota": "Minnesota Twins",    "twins": "Minnesota Twins",
    "yankees": "New York Yankees",     "nyy": "New York Yankees",
    "seattle": "Seattle Mariners",     "mariners": "Seattle Mariners",
    "tampa bay": "Tampa Bay Rays",     "rays": "Tampa Bay Rays",          "tb": "Tampa Bay Rays",
    "athletics": "The Athletics",      "a's": "The Athletics",            "ath": "The Athletics",  "oakland": "The Athletics",
    "texas": "Texas Rangers",          "rangers": "Texas Rangers",
    "toronto": "Toronto Blue Jays",    "blue jays": "Toronto Blue Jays",  "jays": "Toronto Blue Jays",
}
_NL_ALIASES = {
    "arizona": "Arizona Diamondbacks", "diamondbacks": "Arizona Diamondbacks", "d-backs": "Arizona Diamondbacks", "dbacks": "Arizona Diamondbacks",
    "atlanta": "Atlanta Braves",       "braves": "Atlanta Braves",
    "chicago": "Chicago Cubs",         "cubs": "Chicago Cubs",
    "cincinnati": "Cincinnati Reds",   "reds": "Cincinnati Reds",
    "colorado": "Colorado Rockies",    "rockies": "Colorado Rockies",
    "los angeles": "Los Angeles Dodgers", "la": "Los Angeles Dodgers",   "dodgers": "Los Angeles Dodgers",
    "miami": "Miami Marlins",          "marlins": "Miami Marlins",
    "milwaukee": "Milwaukee Brewers",  "brewers": "Milwaukee Brewers",
    "mets": "New York Mets",           "nym": "New York Mets",            "new york mets": "New York Mets",
    "philadelphia": "Philadelphia Phillies", "phillies": "Philadelphia Phillies", "phils": "Philadelphia Phillies",
    "pittsburgh": "Pittsburgh Pirates","pirates": "Pittsburgh Pirates",   "bucs": "Pittsburgh Pirates",
    "san diego": "San Diego Padres",   "padres": "San Diego Padres",      "sd": "San Diego Padres",
    "san francisco": "San Francisco Giants", "giants": "San Francisco Giants", "sf": "San Francisco Giants",
    "st. louis": "St. Louis Cardinals","cardinals": "St. Louis Cardinals", "cards": "St. Louis Cardinals",
    "washington": "Washington Nationals","nationals": "Washington Nationals", "nats": "Washington Nationals",
}


def _qh_route(text: str, league: str) -> tuple[str, str | None]:
    """Returns (stripped_text, canonical_team_name | None). If no prefix
    matches, the original text is returned and the second element is None."""
    if not isinstance(text, str):
        return text, None
    m = QH_TEAM_PREFIX_RE.match(text)
    if not m:
        return text, None
    prefix = m.group(1).strip().lower()
    table = _AL_ALIASES if league == "AL" else _NL_ALIASES
    team = table.get(prefix)
    if not team:
        return text, None
    return text[m.end():].lstrip(), team


def strip_qh_team_prefix(text: str) -> str:
    """Strip a leading team-name prefix without re-routing (used in tests)."""
    if not isinstance(text, str):
        return text
    return QH_TEAM_PREFIX_RE.sub("", text, count=1)


def parse_quick_hits(ws, league: str, team_order: list[str]) -> list[dict]:
    """Quick Hits rows: col 2 = date (datetime), cols 3..(3+N-1) hold per-team notes
    in the SAME column order as the team rows (col 3 = team #1, col 4 = team #2, etc.).
    """
    out = []
    max_row = ws.max_row
    for r in range(QUICK_HITS_START_ROW, max_row + 1):
        date_val = ws.cell(row=r, column=2).value
        if date_val is None:
            continue
        # Some rows might be free-form 'Quick Hits...' summaries in col 1 with date in col 2.
        date_str = None
        if hasattr(date_val, "isoformat"):
            date_str = date_val.date().isoformat() if hasattr(date_val, "date") else date_val.isoformat()
        elif isinstance(date_val, str):
            date_str = date_val.strip()
        if not date_str:
            continue

        entries: dict[str, str] = {}
        # col 1 is sometimes a roll-up "Quick Hits..." string spanning multiple teams.
        rollup = ws.cell(row=r, column=1).value
        if isinstance(rollup, str) and rollup.strip():
            entries["__rollup__"] = rollup.strip()

        for idx, team_name in enumerate(team_order):
            col = 3 + idx
            val = ws.cell(row=r, column=col).value
            if not isinstance(val, str) or not val.strip():
                continue
            stripped, routed_team = _qh_route(val.strip(), league)
            # Re-route to the team named by the blurb's prefix; only if no
            # prior blurb is sitting in that slot for this date. If multiple
            # blurbs target the same team, append.
            target = routed_team or team_name
            if target in entries:
                entries[target] = entries[target].rstrip() + "\n\n" + stripped
            else:
                entries[target] = stripped

        if entries:
            out.append({"date": date_str, "league": league, "entries": entries})
    return out


def parse_workbook(xlsx_path: Path) -> tuple[list[dict], list[dict], dict[str, list[str]]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    teams: list[dict] = []
    quickhits: list[dict] = []
    team_order_by_league: dict[str, list[str]] = {"AL": [], "NL": []}

    for sheet, league in (("AL Chart", "AL"), ("NL Chart", "NL")):
        ws = wb[sheet]
        league_teams = []
        for r in TEAM_ROW_RANGE:
            row = parse_team_row(ws, r, league)
            if row:
                teams.append(row)
                league_teams.append(row["teamName"])
        team_order_by_league[league] = league_teams
        quickhits.extend(parse_quick_hits(ws, league, league_teams))

    # Sort quick hits desc by date
    quickhits.sort(key=lambda q: q["date"], reverse=True)
    return teams, quickhits, team_order_by_league


# -----------------------------------------------------------------------------
# Roster fetching


def normalize_name(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


def fetch_rosters() -> dict:
    """Return {teamId: {team, abbr, pitchers:[{id,name,posAbbr,hand,status}]}}.
    Pulls 40-man rosters from statsapi (rosterType=40Man).
    """
    teams = mlb.teams_mlb_only()
    print(f"[rosters] {len(teams)} MLB orgs found", file=sys.stderr)

    out: dict = {}
    for t in teams:
        tid = int(t["id"])
        team_name = t["name"]
        abbr = t["abbreviation"]
        try:
            roster = mlb.team_roster(tid, roster_type="40Man")
        except Exception as e:
            print(f"[rosters] {team_name}: request failed: {e}", file=sys.stderr)
            continue

        pitchers = []
        for p in roster:
            pos = p.get("position", {})
            if pos.get("abbreviation") != "P":
                continue
            person = p.get("person", {})
            pitchers.append({
                "id": person.get("id"),
                "name": person.get("fullName"),
                "posAbbr": pos.get("abbreviation"),
                "hand": (person.get("pitchHand") or {}).get("code"),
                "status": (p.get("status") or {}).get("description"),
            })
        out[str(tid)] = {
            "teamId": tid,
            "team": team_name,
            "abbr": abbr,
            "league": t.get("leagueName"),
            "division": t.get("divisionName"),
            "pitchers": sorted(pitchers, key=lambda x: (x["name"] or "").lower()),
        }
        print(f"[rosters] {abbr:>3} {team_name}: {len(pitchers)} pitchers", file=sys.stderr)
    return out


def attach_team_ids(teams: list[dict], rosters: dict) -> None:
    """Map team display name -> teamId by exact name match."""
    name_to_id = {info["team"]: int(info["teamId"]) for info in rosters.values()}
    # Special-case "The Athletics" since the chart uses that label (vs MLB's "Athletics" or "Oakland Athletics").
    aliases = {
        "The Athletics": "Athletics",
    }
    for t in teams:
        nm = t["teamName"]
        tid = name_to_id.get(nm) or name_to_id.get(aliases.get(nm, ""))
        if tid is None:
            # Fuzzy fallback
            choice = process.extractOne(nm, list(name_to_id.keys()), scorer=fuzz.WRatio, score_cutoff=80)
            if choice:
                tid = name_to_id[choice[0]]
                print(f"[teamMap] fuzzy '{nm}' -> '{choice[0]}' (score {choice[1]:.0f})", file=sys.stderr)
        t["teamId"] = tid
        if tid is None:
            print(f"[teamMap] UNRESOLVED team '{nm}'", file=sys.stderr)


# -----------------------------------------------------------------------------
# Name -> MLBAMID resolution


def build_name_map(teams: list[dict], rosters: dict, prior_map: dict, *, auto: bool) -> dict:
    """Return updated {originalName: mlbamid|null}.

    For each chip across all teams:
      - if name already resolved in prior_map -> keep
      - else fuzzy-match against that team's pitcher list
      - score >=92 -> auto-accept
      - else interactive prompt (or auto pick best with --auto)
    """
    name_map = dict(prior_map)
    AUTO_THRESHOLD = 92      # silent accept at this score
    PROMPT_THRESHOLD = 85    # in --auto mode, accept above this; below -> unresolved
                             # (interactive mode prompts whenever we're below AUTO_THRESHOLD)

    for t in teams:
        tid = t.get("teamId")
        if tid is None:
            continue
        roster = rosters[str(tid)]
        pitcher_names = [p["name"] for p in roster["pitchers"] if p["name"]]
        norm_to_pitcher = {normalize_name(n): n for n in pitcher_names}

        # Build list of unique chip names for this team (across all roles).
        chip_names: list[str] = []
        for role_chips in t["roles"].values():
            for chip in role_chips:
                if chip["name"] and chip["name"] not in chip_names:
                    chip_names.append(chip["name"])

        for nm in chip_names:
            if nm in name_map and name_map[nm] is not None:
                continue
            if not pitcher_names:
                name_map.setdefault(nm, None)
                continue

            # Try normalized exact
            nrm = normalize_name(nm)
            if nrm in norm_to_pitcher:
                pname = norm_to_pitcher[nrm]
                pid = next(p["id"] for p in roster["pitchers"] if p["name"] == pname)
                name_map[nm] = pid
                continue

            # Fuzzy
            ranked = process.extract(nm, pitcher_names, scorer=fuzz.WRatio, limit=5)
            best = ranked[0] if ranked else None
            if best and best[1] >= AUTO_THRESHOLD:
                pid = next(p["id"] for p in roster["pitchers"] if p["name"] == best[0])
                name_map[nm] = pid
                continue

            if auto:
                if best and best[1] >= PROMPT_THRESHOLD:
                    pid = next(p["id"] for p in roster["pitchers"] if p["name"] == best[0])
                    name_map[nm] = pid
                    print(f"[names] auto: '{nm}' -> '{best[0]}' (score {best[1]:.0f}) on {roster['abbr']}", file=sys.stderr)
                else:
                    name_map[nm] = None
                    print(f"[names] UNRESOLVED '{nm}' on {roster['abbr']} (best: {best})", file=sys.stderr)
                continue

            # Interactive
            print(f"\n[{roster['abbr']}] '{nm}' — choose match:")
            for i, (pname, score, _) in enumerate(ranked, 1):
                print(f"  {i}) {pname:<28}  (score {score:.0f})")
            print(f"  s) skip / leave unresolved")
            print(f"  c) custom name (type a substring to re-search)")
            sel = input("> ").strip().lower()
            if sel == "s" or sel == "":
                name_map[nm] = None
            elif sel == "c":
                q = input("search: ").strip()
                hits = process.extract(q, pitcher_names, scorer=fuzz.WRatio, limit=10)
                for i, (pname, score, _) in enumerate(hits, 1):
                    print(f"  {i}) {pname:<28}  (score {score:.0f})")
                pick = input("> ").strip()
                try:
                    idx = int(pick) - 1
                    pname = hits[idx][0]
                    pid = next(p["id"] for p in roster["pitchers"] if p["name"] == pname)
                    name_map[nm] = pid
                except (ValueError, IndexError):
                    name_map[nm] = None
            else:
                try:
                    idx = int(sel) - 1
                    pname = ranked[idx][0]
                    pid = next(p["id"] for p in roster["pitchers"] if p["name"] == pname)
                    name_map[nm] = pid
                except (ValueError, IndexError):
                    name_map[nm] = None
    return name_map


def attach_chip_ids(teams: list[dict], name_map: dict) -> None:
    for t in teams:
        for role_chips in t["roles"].values():
            for chip in role_chips:
                chip["mlbamid"] = name_map.get(chip["name"])


# -----------------------------------------------------------------------------
# Output


def build_chart(teams: list[dict]) -> dict:
    return {
        "teams": [{
            "teamId": t.get("teamId"),
            "teamName": t["teamName"],
            "league": t["league"],
            "levcon": t["levcon"],
            "roles": t["roles"],
            "notes": t["notes"],
        } for t in teams],
        "roleOrder": ["closer", "stopper", "stealth", "upside", "hybrid"],
        "roleLabels": {
            "closer":  "Closer (highest leverage)",
            "stopper": "Stopper (unless fluid)",
            "stealth": "Stealth Option",
            "upside":  "High Upside Future",
            "hybrid":  "Hybrid (IP+) or RP Pivot",
        },
        # Per-chip color palette. Each label is what shows in the legend.
        "colorMeanings": {
            "Blue":    "Fluid bullpen",
            "Orange":  "Role at risk",
            "Yellow":  "Injury Concern/Shutdown",
            "Magenta": "Workload concern/on leave",
            "Green":   "Upside Reliever/Stash",
        },
        "levconLevels": {
            5: "Clear cut closer",
            4: "Preferred/HLR",
            3: "Shared saves",
            2: "On alert (ratios/role)",
            1: "Committee w/o clarity",
        },
    }


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    ap.add_argument("--out", default=str(DEFAULT_OUT))
    ap.add_argument("--auto", action="store_true", help="auto-accept best name guess (no prompts)")
    ap.add_argument("--skip-rosters", action="store_true", help="reuse existing rosters.json")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"[parse] reading {xlsx_path}", file=sys.stderr)
    teams, quickhits, _team_order = parse_workbook(xlsx_path)
    print(f"[parse] {len(teams)} teams, {len(quickhits)} quick-hit dates", file=sys.stderr)

    rosters_path = out_dir / "rosters.json"
    if args.skip_rosters and rosters_path.exists():
        print(f"[rosters] reusing {rosters_path}", file=sys.stderr)
        rosters = json.loads(rosters_path.read_text())
    else:
        rosters = fetch_rosters()
        rosters_path.write_text(json.dumps(rosters, indent=2, ensure_ascii=False))
        print(f"[rosters] wrote {rosters_path}", file=sys.stderr)

    attach_team_ids(teams, rosters)

    name_map_path = out_dir / "name_map.json"
    prior_map = json.loads(name_map_path.read_text()) if name_map_path.exists() else {}
    print(f"[names] resolving against rosters (auto={args.auto})", file=sys.stderr)
    name_map = build_name_map(teams, rosters, prior_map, auto=args.auto)
    name_map_path.write_text(json.dumps(name_map, indent=2, ensure_ascii=False))
    print(f"[names] wrote {name_map_path} ({sum(1 for v in name_map.values() if v)} resolved / {len(name_map)} total)", file=sys.stderr)

    attach_chip_ids(teams, name_map)

    chart = build_chart(teams)
    chart_path = out_dir / "chart.json"
    chart_path.write_text(json.dumps(chart, indent=2, ensure_ascii=False))
    print(f"[chart] wrote {chart_path}", file=sys.stderr)

    # Map quickhits team-name keys -> teamIds where possible
    name_to_id = {t["teamName"]: t.get("teamId") for t in teams}
    qh_out = []
    for q in quickhits:
        new_entries = {}
        for k, v in q["entries"].items():
            if k == "__rollup__":
                new_entries[k] = v
                continue
            tid = name_to_id.get(k)
            new_entries[str(tid) if tid else k] = v
        qh_out.append({"date": q["date"], "league": q["league"], "entries": new_entries})

    qh_path = out_dir / "quickhits.json"
    qh_path.write_text(json.dumps(qh_out, indent=2, ensure_ascii=False))
    print(f"[quickhits] wrote {qh_path} ({len(qh_out)} dates)", file=sys.stderr)


if __name__ == "__main__":
    main()
