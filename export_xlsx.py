"""Round-trip data/chart.json + data/quickhits.json back to .xlsx.

Strategy: open the existing workbook (preserves formulas, formatting, fills, fonts,
column widths, the legend column, etc.) and overwrite ONLY the cells the webapp owns:
  - col 1 (color tag) on team rows -- only writes when value is in KNOWN_COLORS
  - col 2 (LEVCON rating)
  - cols 4..8 (the five role columns), joined as 'Name (tag) share/Other Name'
  - col 10 (notes)
  - Quick Hits rows (rows 20+) get rebuilt from quickhits.json

Does NOT touch the legend cells in column 1 on non-team rows.

Usage:
  python3 export_xlsx.py                              # writes Closer Charts (export).xlsx next to source
  python3 export_xlsx.py --out path/to/file.xlsx
  python3 export_xlsx.py --in-place                   # overwrites source workbook
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
SRC_XLSX = ROOT / "Copy of Closer Charts.xlsx"
DATA_DIR = ROOT / "data"

ROLE_COLS = {
    "closer": 4, "stopper": 5, "stealth": 6, "upside": 7, "hybrid": 8,
}
LEVCON_COL = 2
TEAM_COL = 3
COLOR_COL = 1
NOTES_COL = 10
TEAM_ROWS = list(range(4, 19))  # 4..18 inclusive
QH_START = 20


def chip_to_text(chip: dict) -> str:
    """Render a chip back to a single role-cell token.

    Annotation syntax (matches build_data.split_chip_text):
      Name [Blue] (paternity) {HLR}
       ^      ^         ^         ^
       |   color    statusTag   other

    Usage tags are NOT included — they're derived live from gameLog every time
    Refresh Stats runs.
    """
    parts = [chip.get("name", "").strip()]
    color = chip.get("color")
    status = chip.get("statusTag")
    other = chip.get("other")
    if color:
        parts.append(f"[{color}]")
    if status:
        parts.append(f"({status})")
    if other:
        parts.append("{" + other + "}")
    return " ".join(p for p in parts if p)


def role_cell_text(chips: list[dict]) -> str:
    return "/".join(chip_to_text(c) for c in chips if c.get("name"))


def write_team_row(ws, row: int, team: dict) -> None:
    # Column 1 (team color) is no longer owned by the webapp -- the chart now
    # uses per-chip colors instead of per-team. We deliberately leave col 1
    # alone so existing legend cells / formatting in the source workbook are
    # preserved on re-export.

    # LEVCON rating
    levcon = team.get("levcon")
    ws.cell(row=row, column=LEVCON_COL).value = levcon if levcon is not None else None

    # Roles
    for role, col in ROLE_COLS.items():
        chips = team.get("roles", {}).get(role, [])
        ws.cell(row=row, column=col).value = role_cell_text(chips) or None

    # Notes
    notes = team.get("notes") or ""
    ws.cell(row=row, column=NOTES_COL).value = notes if notes.strip() else None


def clear_quickhits(ws, n_team_cols: int) -> None:
    """Wipe rows from QH_START to ws.max_row across cols 1..(2+n_team_cols)."""
    for r in range(QH_START, ws.max_row + 1):
        for c in range(1, 2 + n_team_cols + 1):
            ws.cell(row=r, column=c).value = None


def write_quickhits(ws, league: str, quickhits: list[dict], team_order: list[dict]) -> None:
    # team_order = list of {teamId, teamName} in column order matching cols 3+
    n = len(team_order)
    clear_quickhits(ws, n)

    # Filter and sort: descending by date for the league
    rows = [q for q in quickhits if q.get("league") == league]
    rows.sort(key=lambda q: q["date"], reverse=True)

    for i, q in enumerate(rows):
        r = QH_START + i
        # rollup col 1
        ru = q.get("entries", {}).get("__rollup__")
        if ru:
            ws.cell(row=r, column=1).value = ru
        # date col 2
        try:
            ws.cell(row=r, column=2).value = datetime.fromisoformat(q["date"]).date()
        except (ValueError, TypeError):
            ws.cell(row=r, column=2).value = q["date"]
        # per-team
        for idx, t in enumerate(team_order):
            col = 3 + idx
            tid = str(t["teamId"]) if t.get("teamId") is not None else None
            txt = (
                q["entries"].get(tid)
                or q["entries"].get(t["teamName"])
                or ""
            )
            if txt:
                ws.cell(row=r, column=col).value = txt


def export(src_path: Path, out_path: Path) -> None:
    chart = json.loads((DATA_DIR / "chart.json").read_text())
    quickhits = json.loads((DATA_DIR / "quickhits.json").read_text())

    wb = openpyxl.load_workbook(src_path)

    # Build per-league team row lookup based on the chart's order
    by_league: dict[str, list[dict]] = {"AL": [], "NL": []}
    for t in chart["teams"]:
        by_league[t["league"]].append(t)

    for sheet, league in (("AL Chart", "AL"), ("NL Chart", "NL")):
        ws = wb[sheet]
        teams = by_league[league]
        if len(teams) > len(TEAM_ROWS):
            print(f"[warn] {league}: {len(teams)} teams but only {len(TEAM_ROWS)} rows allocated", file=sys.stderr)
        for row, team in zip(TEAM_ROWS, teams):
            write_team_row(ws, row, team)
        write_quickhits(ws, league, quickhits, teams)

    wb.save(out_path)
    print(f"[export] wrote {out_path}", file=sys.stderr)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", default=str(SRC_XLSX), help="source .xlsx (formatting template)")
    ap.add_argument("--out", default=str(SRC_XLSX.with_name("Closer Charts (export).xlsx")))
    ap.add_argument("--in-place", action="store_true", help="overwrite --src instead of writing --out")
    args = ap.parse_args()

    src = Path(args.src)
    out = Path(args.src) if args.in_place else Path(args.out)
    export(src, out)


if __name__ == "__main__":
    main()
