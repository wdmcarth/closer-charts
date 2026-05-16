"""One-off importer: merge an updated 'Closer Charts.xlsx' on disk into the
live data/chart.json. PRESERVES per-chip color tags, status tags, other tags,
noMagenta dismissals, and MLBAMID bindings whenever a chip's name still
matches an existing chip in the same team+role. Replaces:

  - team.levcon  (LEVCON rating)
  - team.notes   (right-hand injury/availability column)
  - team.roles.*  (chip list per role)

Usage:
  python3 update_from_xlsx.py [--xlsx PATH]   default: ./Closer Charts.xlsx
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl
from rapidfuzz import fuzz, process

ROOT = Path(__file__).resolve().parent
DATA = ROOT / "data"
DEFAULT_XLSX = ROOT / "Closer Charts.xlsx"

# Same column layout as build_data.py
ROLE_COLS = {4: "closer", 5: "stopper", 6: "stealth", 7: "upside", 8: "hybrid"}
TEAM_COL, LEVCON_COL, NOTES_COL = 3, 2, 10
TEAM_ROWS = range(4, 19)

# Parse-time regexes — same idea as build_data but tightened to strip the
# usage-tag noise the user types into the xlsx (b2b, 1.1 IP, 30 P, etc.)
# so the chip name matches a real roster pitcher.
SHARE_RE = re.compile(r"\b\d+\s*-\s*of\s*-\s*\d+\b", re.IGNORECASE)
STREAK_RE = re.compile(r"\bb2b\b|\b[34]-in-a-row\b", re.IGNORECASE)
VOLUME_RE = re.compile(r"\b\d+(?:\.\d+)?\s*IP\b|\b\d+\s*(?:np|P)\b", re.IGNORECASE)
COLOR_RE = re.compile(r"\[([^\]]+)\]")
OTHER_RE = re.compile(r"\{([^}]+)\}")
PAREN_RE = re.compile(r"\(([^)]+)\)")


def split_chip_text(raw: str) -> list[dict]:
    """Split a role-cell string into chip dicts. Mirrors build_data but
    additionally strips streak/volume usage annotations the user types inline."""
    if not raw:
        return []
    chips: list[dict] = []
    for piece in re.split(r"\s*/\s*", str(raw).strip()):
        piece = piece.strip()
        if not piece:
            continue

        # Strip usage-tag noise (these are auto-derived; never persist).
        piece = SHARE_RE.sub("", piece)
        piece = STREAK_RE.sub("", piece)
        piece = VOLUME_RE.sub("", piece)

        # Drop bracketed color tags — we'll preserve any existing chart-side
        # colors via the merge step, not from the xlsx.
        piece = COLOR_RE.sub("", piece)

        # Parenthetical → status tag (only if it looks like one — IL, paternity, etc.)
        status_vals: list[str] = []
        for pm in PAREN_RE.finditer(piece):
            status_vals.append(pm.group(1).strip())
        piece = PAREN_RE.sub("", piece)

        # Curly → other note
        other_vals: list[str] = []
        for om in OTHER_RE.finditer(piece):
            other_vals.append(om.group(1).strip())
        piece = OTHER_RE.sub("", piece)

        # Clean up: drop semicolons / commas left over from usage-tag stripping
        piece = re.sub(r"[;,]+", " ", piece)
        name = re.sub(r"\s+", " ", piece).strip()
        if not name:
            continue
        chips.append({
            "name": name,
            "statusTag": status_vals[0] if status_vals else None,
            "other": other_vals[0] if other_vals else None,
        })
    return chips


def parse_xlsx(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    out: list[dict] = []
    for sheet, league in (("AL Chart", "AL"), ("NL Chart", "NL")):
        ws = wb[sheet]
        for r in TEAM_ROWS:
            team_name = ws.cell(row=r, column=TEAM_COL).value
            if not isinstance(team_name, str) or not team_name.strip():
                continue

            lev = ws.cell(row=r, column=LEVCON_COL).value
            try:
                lev = int(lev) if lev is not None else None
            except (TypeError, ValueError):
                lev = None

            notes = ws.cell(row=r, column=NOTES_COL).value
            notes = notes.strip() if isinstance(notes, str) else ""

            roles: dict[str, list[dict]] = {}
            for col, role_key in ROLE_COLS.items():
                cell_val = ws.cell(row=r, column=col).value
                roles[role_key] = split_chip_text(cell_val if isinstance(cell_val, str) else "")

            out.append({
                "teamName": team_name.strip(),
                "league": league,
                "levcon": lev,
                "notes": notes,
                "roles": roles,
            })
    return out


def merge_chip(new_chip: dict, old_chip: dict | None) -> dict:
    """Return a chip dict with the new name + parsed status/other, but
    preserving the OLD chip's user-set metadata (colors, noMagenta, mlbamid)
    when names match. If no old_chip, returns the new with empty colors."""
    if not old_chip:
        return {
            "name": new_chip["name"],
            "mlbamid": None,                      # will be resolved later
            "colors": [],
            "statusTag": new_chip.get("statusTag"),
            "other": new_chip.get("other"),
        }
    # Reuse old chip's user-set metadata where it exists.
    # Legacy compat: old chip may have chip.color (string) instead of colors.
    colors = old_chip.get("colors")
    if not isinstance(colors, list):
        colors = [old_chip["color"]] if old_chip.get("color") else []
    merged = {
        "name": new_chip["name"],
        "mlbamid": old_chip.get("mlbamid"),
        "colors": colors,
        # Prefer the new statusTag/other only if old didn't have one; respect
        # the user's manual edits otherwise.
        "statusTag": old_chip.get("statusTag") or new_chip.get("statusTag"),
        "other": old_chip.get("other") or new_chip.get("other"),
    }
    if old_chip.get("noMagenta"):
        merged["noMagenta"] = True
    return merged


def merge_teams(new_teams: list[dict], chart: dict) -> dict:
    existing_by_name = {t["teamName"]: t for t in chart["teams"]}
    for new_team in new_teams:
        old_team = existing_by_name.get(new_team["teamName"])
        if not old_team:
            print(f"[update] WARN: no existing team for {new_team['teamName']!r}",
                  file=sys.stderr)
            continue
        old_team["levcon"] = new_team["levcon"]
        old_team["notes"] = new_team["notes"]
        for role, new_chips in new_team["roles"].items():
            old_chips = old_team.get("roles", {}).get(role, []) or []
            old_by_name = {(c.get("name") or "").lower(): c for c in old_chips}
            merged: list[dict] = []
            for nc in new_chips:
                key = (nc.get("name") or "").lower()
                merged.append(merge_chip(nc, old_by_name.get(key)))
            old_team.setdefault("roles", {})[role] = merged
    return chart


def resolve_unbound(chart: dict, rosters: dict, name_map: dict) -> dict:
    """For each chip with mlbamid=None, try the persisted name_map first,
    then a per-team fuzzy match against that team's 40-man roster."""
    for team in chart["teams"]:
        tid = team.get("teamId")
        if not tid:
            continue
        roster = rosters.get(str(tid))
        if not roster:
            continue
        pitcher_names = [p["name"] for p in roster.get("pitchers", []) if p.get("name")]
        for role in team.get("roles", {}):
            for chip in team["roles"][role]:
                if chip.get("mlbamid"):
                    continue
                # 1) name_map hit
                if chip["name"] in name_map and name_map[chip["name"]]:
                    chip["mlbamid"] = name_map[chip["name"]]
                    continue
                # 2) fuzzy match on the team's roster
                if not pitcher_names:
                    continue
                best = process.extractOne(
                    chip["name"], pitcher_names, scorer=fuzz.WRatio, score_cutoff=85
                )
                if best:
                    pname = best[0]
                    pid = next(p["id"] for p in roster["pitchers"] if p["name"] == pname)
                    chip["mlbamid"] = pid
                    name_map[chip["name"]] = pid
    return chart


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"[update] missing {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    print(f"[update] parsing {xlsx_path.name}", file=sys.stderr)
    new_teams = parse_xlsx(xlsx_path)
    print(f"[update] {len(new_teams)} teams parsed", file=sys.stderr)

    chart = json.loads((DATA / "chart.json").read_text())
    chart = merge_teams(new_teams, chart)

    rosters_path = DATA / "rosters.json"
    name_map_path = DATA / "name_map.json"
    if rosters_path.exists():
        rosters = json.loads(rosters_path.read_text())
        name_map = json.loads(name_map_path.read_text()) if name_map_path.exists() else {}
        chart = resolve_unbound(chart, rosters, name_map)
        name_map_path.write_text(json.dumps(name_map, indent=2, ensure_ascii=False))

    (DATA / "chart.json").write_text(json.dumps(chart, indent=2, ensure_ascii=False))

    # Report
    teams = chart["teams"]
    total = sum(len(t.get("roles", {}).get(r, [])) for t in teams for r in t.get("roles", {}))
    resolved = sum(
        1 for t in teams for r in t.get("roles", {})
        for c in t["roles"][r] if c.get("mlbamid")
    )
    preserved_colors = sum(
        1 for t in teams for r in t.get("roles", {})
        for c in t["roles"][r] if (c.get("colors") or [])
    )
    print(f"[update] wrote data/chart.json — {total} chips, "
          f"{resolved} resolved, {preserved_colors} kept user-set colors",
          file=sys.stderr)


if __name__ == "__main__":
    main()
